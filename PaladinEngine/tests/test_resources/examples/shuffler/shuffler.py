"""The shuffleinator:
    this is an app created to check the calibration of the catch rate in 'pokemon shuffle'
    it does that by collecting data from users at Ofek's friend group
    the collected data is entered to an Excel sheet and graphed where the X-axis is the
    percents given by the app and the Y-axis represents our measured percents
    the other purpose of the app is to give Ofek an opportunity to learn python in a fun way
    which will cover most of the topics he needs for the Goomy app"""

import matplotlib
import matplotlib.pyplot as plt

# this is the program that know to write the data and extract it in the correct way from the Excel file
import openpyxl
import xlsxwriter

PATH = "/Users/orenafek/Projects/Paladin/PaladinEngine/PaladinEngine/tests/test_resources/examples/shuffler/ShuffleData.xlsx"

def write_to_cell(percent, result, path=PATH):
    """
    the command that write into the Excel file the information received from the user
    work's by loading the file, increasing the counter of the percent given by one and entering the amount caught,
        then divide the amount caught by the amount of tries.
    :param percent: int, 0<x<=100,
    :param result: int, 0,1
    :param path: of the Excel file
    :return: does not return, write into the file
    """
    file = openpyxl.load_workbook(path)
    sheet = file[file.sheetnames[0]]
    col_letter = xlsxwriter.utility.xl_col_to_name(percent) # translate num to Excel column letter
    current_index_location = col_letter + "3"
    current_index = int(sheet[current_index_location].value)
    new_index = current_index + 1
    sheet[current_index_location].value = new_index
    writing_index_location = col_letter + "4"
    writing_index = int(sheet[writing_index_location].value)
    if result == 1:
        sheet[writing_index_location].value = writing_index + 1
        research_index = col_letter + "2"
        sheet[research_index].value = (writing_index + 1) / new_index
    else:
        research_index = col_letter + "2"
        sheet[research_index].value = writing_index / new_index
    file.save(path)


def check_percent(percent):
    if percent.isnumeric():
        if 0 < int(percent) <= 100:
            return True
    return False


def export_plot_data(path=PATH):
    file = openpyxl.load_workbook(path)
    sheet = file[file.sheetnames[0]]
    percents = []
    numeric_data = []
    for i in range(1, 101):
        column_num = xlsxwriter.utility.xl_col_to_name(i)
        cell_check = sheet[column_num + "2"].value
        if cell_check is not None:
            percents.append(i)
            numeric_data.append(100 * float(cell_check))
    file.save(path)
    return percents, numeric_data


matplotlib.use("TkAgg")


def shuffle_plot(percents, data, r_square=1):
    plt.plot(percents, data, 'bo')
    plt.plot(range(101), range(101), 'g--')
    plt.title("The ShuffleInator R-Square=" + str(r_square))
    plt.xlabel("Catch percents[%]")
    plt.ylabel("Catch rate[%]")
    plt.xlim([0, 100])
    plt.ylim([0, 100])
    plt.show()



def main():
    Flag1 = True
    while Flag1:
        choice = input("what do you want to do?\nenter data - 1    plot - 2\nexit - 3\n")
        if choice == "1":
            Flag2 = True
            while Flag2:
                percent = input("how many percents did the game said the catch is?\n")
                if check_percent(percent) is True:
                    result = ""
                    while result not in ["0", "1"]:
                        result = input("did you catch it?(1 for catch 0 for no catch)\n")
                    write_to_cell(int(percent), int(result))
                    run_again = input("enter 1 for another entry?\n")
                    if run_again != "1":
                        Flag2 = False
                else:
                    print("please enter a number which is 0 < num <= 100")
        elif choice == "2":
            percents, numeric_data = export_plot_data()
            shuffle_plot(percents, numeric_data)
        elif choice == "3":
            Flag1 = False


main()

